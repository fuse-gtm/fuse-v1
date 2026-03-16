#!/usr/bin/env python3

import csv
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
INPUT_CSV = ROOT / "output" / "spreadsheet" / "upstream_commit_review_draft.csv"
FINAL_CSV = ROOT / "output" / "spreadsheet" / "upstream_commit_review_final.csv"
FINAL_XLSX = ROOT / "output" / "spreadsheet" / "upstream_commit_review_final.xlsx"


OVERRIDES = {
    "2e9624858c": ("no", "dev-mode only", "dev_only"),
    "f09a9cc25a": ("no", "cosmetic CSS tweak", "cosmetic"),
    "999dcd4468": ("no", "minor admin UI sizing", "cosmetic"),
    "72086fe111": ("no", "dependency bump only", "dependency"),
    "6a2e0182ab": ("no", "dependency bump only", "dependency"),
    "bfa50f566e": ("no", "dependency bump only", "dependency"),
    "57499342f1": ("no", "widget layout internal tweak", "noncore_feature"),
    "1decd40eea": ("yes", "server perf fix", "perf_fix"),
    "1c898f36d6": ("no", "dark-mode cosmetic", "cosmetic"),
    "73268535dc": ("yes", "query/filter correctness", "core_fix"),
    "66d93c4d28": ("no", "app-dev CLI only", "platform"),
    "5b28e59ca7": ("no", "navigation feature overhaul", "feature"),
    "a9696705c1": ("no", "widget behavior change", "ux_tweak"),
    "2c5af2654d": ("no", "feature-flag plumbing", "noncore_feature"),
    "f0c83434a7": ("yes", "production safety default", "security"),
    "2af3121c51": ("yes", "dashboard creation bug", "core_fix"),
    "2a6fcfcfb3": ("no", "side-panel framework feature", "feature"),
    "5f558e5539": ("no", "dev-only enterprise key path", "dev_only"),
    "f3e0c12ce6": ("no", "app install platform fix", "platform"),
    "0379aea0b1": ("no", "migration tuning only", "migration"),
    "95a35f8a1d": ("no", "new OAuth capability", "feature"),
    "a07337fea0": ("no", "MCP/platform response change", "platform"),
    "26f0a416a1": ("no", "too large for safe cherry-pick; manual port later", "manual_port"),
    "c571473d67": ("no", "superseded by later SVGO lockfile fix", "superseded"),
    "93bd4960e3": ("no", "side-column styling", "cosmetic"),
    "1735c7527c": ("no", "2FA layout polish", "cosmetic"),
    "1be87eb97b": ("no", "dead code cleanup", "refactor"),
    "731e297147": ("no", "SDK CLI auth feature", "platform"),
    "dedcf4e9b9": ("no", "command-menu feature", "feature"),
    "87efaf2ff8": ("no", "export command feature", "feature"),
    "b6c62b3812": ("yes", "auth token renewal fix", "security"),
    "c6f11d8adb": ("no", "driver refactor", "refactor"),
    "ffb02a6878": ("no", "workflow metrics only", "workflow_feature"),
    "9a850e2241": ("no", "billing recovery logic", "enterprise"),
    "7bde8a4dfa": ("no", "standard-view migration", "migration"),
    "c9ca4dfa14": ("no", "AI catalog sync plumbing", "platform"),
    "b813d64324": ("no", "AI catalog sync data", "platform"),
    "24055527c8": ("no", "wording tweak only", "cosmetic"),
    "fd044ba9a2": ("no", "AI catalog sync data", "platform"),
    "e0b36c3931": ("no", "AI config formatting", "noncore_feature"),
    "d3c7b0131d": ("no", "lambda-driver specific", "platform"),
    "341c13bf32": ("no", "docs only", "docs"),
    "d5b41b2801": ("no", "auth utility refactor", "refactor"),
    "766f956a15": ("no", "icon color polish", "cosmetic"),
    "b642d1b114": ("no", "sidebar animation only", "cosmetic"),
    "578d990b9c": ("no", "AI composer redesign", "cosmetic"),
    "cfefe9273b": ("no", "merge-records feature", "feature"),
    "82611de9b6": ("no", "connected-account follow-up", "ambiguous_no"),
    "b171a23216": ("no", "AI permissions styling", "cosmetic"),
    "695518a15e": ("no", "cosmetic chip sizing", "cosmetic"),
    "6360fb3bce": ("no", "AI catalog sync data", "platform"),
    "f651413297": ("no", "AI catalog sync data", "platform"),
    "cd2e08b912": ("no", "multi-workspace limit feature", "feature"),
    "58189e1c05": ("no", "AI catalog sync data", "platform"),
    "2fccd29ec6": ("no", "messaging cleanup", "refactor"),
    "2263e14394": ("no", "favorites padding polish", "cosmetic"),
    "08f019e9c9": ("no", "AI redesign only", "cosmetic"),
    "887e0283c5": ("no", "direct-execution follow-up", "feature"),
    "20ac5b7e84": ("no", "field-widget edit feature", "feature"),
    "691c86a4a7": ("no", "front-component platform fix", "platform"),
    "223943550c": ("no", "AI/code-interpreter polish", "cosmetic"),
    "c854d28d60": ("no", "AI logo theming", "cosmetic"),
    "a303d9ca1b": ("no", "direct-execution feature", "feature"),
    "b8e7179a85": ("yes", "missing DB index", "perf_fix"),
    "6f89098340": ("no", "hover overflow cosmetic", "cosmetic"),
    "cd23a2bc80": ("no", "upgrade-command cleanup", "refactor"),
    "aadbf67ec8": ("no", "editor-only behavior", "ux_tweak"),
    "779e613df3": ("no", "mostly design tweak", "cosmetic"),
    "a53b9cb68a": ("no", "SDK generation job", "platform"),
    "a2533baa66": ("no", "obscure local-driver fix", "ambiguous_no"),
    "12b031b67d": ("no", "AI catalog sync data", "platform"),
    "8543576dae": ("no", "command-menu feature", "feature"),
    "4ba2f3b184": ("no", "event-stream cache perf", "ambiguous_no"),
    "f8c3960cf2": ("no", "empty-tab UX tweak", "ux_tweak"),
    "ed912ec548": ("no", "AI catalog sync data", "platform"),
    "8d61bb9ae6": ("no", "messageFolder schema migration", "migration"),
    "8702300b07": ("no", "apps-platform fix", "platform"),
    "137e068ced": ("no", "lambda build target", "platform"),
    "cb85a1b5a3": ("no", "hover-only UI polish", "cosmetic"),
    "6c128a35dd": ("no", "calendar text polish", "cosmetic"),
    "ec459d8dc8": ("no", "sidebar icon polish", "cosmetic"),
    "2317a701bd": ("no", "arrow icon polish", "cosmetic"),
    "191a277ddf": ("yes", "permission cache invalidation", "core_fix"),
    "a51b5ed589": ("no", "usage-chart billing fix", "enterprise"),
    "b1a7155431": ("no", "layout edit-mode guard", "ux_tweak"),
    "67d34be7c8": ("no", "feature-flag follow-up", "ambiguous_no"),
    "2d6c8be7df": ("no", "app-platform search fix", "platform"),
    "baa4fda3d2": ("no", "workflow seed logic", "workflow_feature"),
    "f01bcf60ee": ("no", "AI markdown rendering", "cosmetic"),
    "1b14e7e1f1": ("yes", "server dependency fix", "core_fix"),
}


def apply_overrides(rows):
    for row in rows:
        override = OVERRIDES.get(row["short_sha"])
        if override is None:
            continue
        recommendation, rationale, tag = override
        row["final_recommendation"] = recommendation
        row["final_rationale"] = rationale
        row["final_tag"] = tag
    return rows


def write_csv(rows):
    fieldnames = list(rows[0].keys())
    with FINAL_CSV.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Cherry Pick Review"
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 8,
        "B": 12,
        "C": 14,
        "D": 42,
        "E": 16,
        "F": 70,
        "G": 34,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 18,
        "O": 28,
        "P": 18,
        "Q": 18,
        "R": 28,
        "S": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for idx, header in enumerate(headers, start=1):
        if header in {"subject", "top_dirs", "draft_rationale", "final_rationale", "commit_url"}:
            for cell in ws[get_column_letter(idx)]:
                cell.alignment = Alignment(wrap_text=True)

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    yes_count = sum(1 for row in rows if row["final_recommendation"] == "yes")
    no_count = sum(1 for row in rows if row["final_recommendation"] == "no")
    summary_rows = [
        ("Reviewed commits", len(rows)),
        ("Recommend cherry-pick", yes_count),
        ("Recommend skip", no_count),
        ("Reviewed against", "origin/main..upstream/main"),
    ]
    for item in summary_rows:
        summary.append(list(item))
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 32

    wb.save(FINAL_XLSX)
    return True


def main():
    rows = list(csv.DictReader(INPUT_CSV.open()))
    for row in rows:
        row["final_recommendation"] = row["draft_recommendation"]
        row["final_rationale"] = row["draft_rationale"]
        row["final_tag"] = row["draft_tag"]

    rows = apply_overrides(rows)
    write_csv(rows)
    wrote_xlsx = write_xlsx(rows)
    print(FINAL_CSV)
    print(f"wrote {len(rows)} rows")
    if wrote_xlsx:
        print(FINAL_XLSX)
    else:
        print("openpyxl not available; xlsx skipped")


if __name__ == "__main__":
    main()
