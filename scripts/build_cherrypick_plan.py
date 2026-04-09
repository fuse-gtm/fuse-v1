#!/usr/bin/env python3

import csv
import re
from collections import Counter, defaultdict
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
INPUT_CSV = ROOT / "output" / "spreadsheet" / "upstream_commit_review_final.csv"
PLAN_CSV = ROOT / "output" / "spreadsheet" / "upstream_commit_plan.csv"
PLAN_XLSX = ROOT / "output" / "spreadsheet" / "upstream_commit_plan.xlsx"


BATCH_INFO = {
    "1A": {
        "wave": 1,
        "wave_name": "Safety First",
        "batch_name": "Auth and Runtime Hardening",
        "goal": "Apply security fixes that reduce real production risk.",
        "validation": "Smoke auth, file serving, OAuth, refresh-token renewal, and CalDAV/IMAP flows.",
        "chunk_size": 8,
    },
    "1B": {
        "wave": 1,
        "wave_name": "Safety First",
        "batch_name": "Dependency Vulnerabilities",
        "goal": "Close known parser, XSS, and path-traversal issues with isolated dependency fixes.",
        "validation": "Apply this batch together, regenerate yarn.lock with install, then run build and a short auth/file smoke test.",
        "chunk_size": 8,
    },
    "2A": {
        "wave": 2,
        "wave_name": "Backend Correctness",
        "batch_name": "Server and Runtime Fixes",
        "goal": "Pull backend fixes that improve correctness without changing product scope.",
        "validation": "Smoke server boot, uploads, object metadata, permissions, and export paths.",
        "chunk_size": 10,
    },
    "2B": {
        "wave": 2,
        "wave_name": "Backend Correctness",
        "batch_name": "Data, Query, and Cache Fixes",
        "goal": "Fix query semantics, data integrity, cache invalidation, and backend performance.",
        "validation": "Exercise filters, relations, views, permissions, caching, and CSV/export flows.",
        "chunk_size": 10,
    },
    "3A": {
        "wave": 3,
        "wave_name": "Frontend Correctness",
        "batch_name": "Navigation and Record UI",
        "goal": "Stabilize navigation, board, dashboard, and record-page behavior.",
        "validation": "Smoke nav/sidebar drag-drop, boards, dashboards, settings, and dropdown flows.",
        "chunk_size": 12,
    },
    "3B": {
        "wave": 3,
        "wave_name": "Frontend Correctness",
        "batch_name": "Forms, Inputs, and Editors",
        "goal": "Fix form, picker, upload, and editor bugs that affect basic usage.",
        "validation": "Smoke onboarding, form fields, date pickers, uploads, and text inputs.",
        "chunk_size": 10,
    },
    "3C": {
        "wave": 3,
        "wave_name": "Frontend Correctness",
        "batch_name": "AI and Chat Surface Fixes",
        "goal": "Pull small AI/chat bug fixes without importing larger AI product changes.",
        "validation": "Smoke AI chat compose, message preview, and thread rendering.",
        "chunk_size": 8,
    },
    "3D": {
        "wave": 3,
        "wave_name": "Frontend Correctness",
        "batch_name": "Misc Product Bugs",
        "goal": "Catch remaining user-facing bugs that are neither workflow nor security issues.",
        "validation": "Run focused smoke tests around the touched UI before moving on.",
        "chunk_size": 10,
    },
    "4A": {
        "wave": 4,
        "wave_name": "Workflow Stability",
        "batch_name": "Workflow Reliability",
        "goal": "Fix workflow builder and runtime bugs after the core app is stable.",
        "validation": "Create, edit, and run workflows; confirm dates, titles, and run pages behave.",
        "chunk_size": 8,
    },
    "5A": {
        "wave": 5,
        "wave_name": "Optional Changes",
        "batch_name": "Optional Product Behavior",
        "goal": "Pull lower-urgency behavior improvements only after the high-value fixes land cleanly.",
        "validation": "Smoke the specific UX after each chunk; these are nice-to-have, not urgent.",
        "chunk_size": 8,
    },
}


BATCH_ORDER = ["1A", "1B", "2A", "2B", "3A", "3B", "3C", "3D", "4A", "5A"]


BATCH_OVERRIDES = {
    "dd2a09576b": "3B",
    "3747af4b5a": "3B",
    "c59f420d21": "5A",
    "a3c392ce8b": "5A",
    "262f9f5fe1": "5A",
    "dfd28f5b4a": "5A",
    "0641e07ca6": "2B",
    "3204175065": "2A",
    "523289efad": "2B",
    "4c6e102493": "5A",
    "f47608de07": "5A",
    "4dfd6426c2": "5A",
    "77315b7f6e": "5A",
    "660536d6bb": "3B",
    "ee479001a1": "3A",
    "8eff9efab2": "3A",
    "2af3121c51": "3A",
    "59029a0035": "3A",
}


CHUNK_OVERRIDES = {
    "cd651f57cb": "2B-03",
    "d69e4d7008": "2B-03",
    "0641e07ca6": "2B-03",
    "e552704201": "2B-02",
    "fb5b68f1d8": "2B-02",
    "a8c445a1c2": "2B-02",
    "394a3cef15": "2B-02",
    "8005b35b56": "2B-02",
    "217adc8d17": "2B-02",
    "246afe0f2a": "2B-02",
}


WATCHLIST = [
    {
        "short_sha": "0379aea0b1",
        "subject": "fix: split tsvector migration, add configurable DB timeout, reorder 1.19 commands (#18614)",
        "why_watch": "Potential prerequisite if later schema sync hits migration timeout or ordering problems.",
    },
    {
        "short_sha": "58336fb70f",
        "subject": "fix: navigation menu item type backfill and frontend loading (#18730)",
        "why_watch": "Potential prerequisite if nav-menu schema state is inconsistent after later pulls.",
    },
    {
        "short_sha": "34b3158308",
        "subject": "fix: backfill missing ids on SELECT/MULTI_SELECT field metadata options (#18797)",
        "why_watch": "Potential prerequisite if later field metadata assumes option IDs already exist.",
    },
]


SECURITY_RX = re.compile(
    r"security|oauth|token|password reset|ssrf|xss|blocklist|server logs leak|security headers|vulnerab|cve|path traversal|stored xss|revocation|production build|smtp command injection|cross-site scripting|denial of service",
    re.I,
)
SECURITY_DEP_RX = re.compile(
    r"svgo|yauzl|brace-expansion|node-tar|nodemailer|handlebars|multer|dompurify|mailparser|expr-eval",
    re.I,
)
WORKFLOW_RX = re.compile(r"workflow|flow tab|agent nodes|cron|iterator", re.I)
BACKEND_DATA_RX = re.compile(
    r"query|filter|uuid|relation|cache|aggregate|index|transaction|schema-version|left join|morph|dual-write|app feedbacks|lambda|messagefolder|viewfieldgroup|load more|cache invalidation|transient error|participants|option id|required|health indicator|null-equivalent|log_levels|csv|s3 driver|imap-smtp-caldav|server perf|permission cache|graph 400|metadata version|throttle|subdomain",
    re.I,
)
FRONT_NAV_RX = re.compile(
    r"\bnavigation\b|\bsidebar\b|\bnavbar\b|side panel|\btab\b|\bboard\b|\bdashboard\b|\bcalendar\b|\bbreadcrumb\b|\bdropdown\b|record index|\bwidget\b|\bview\b|\bfolder\b|menu item|system objects|permission settings|timeline|kanban|workspace dropdown|search fallback",
    re.I,
)
FRONT_FORM_RX = re.compile(
    r"\binput\b|\bform\b|\bpicker\b|\bcurrency\b|\bupload\b|\bonboarding\b|json\.parse|\bescape\b|task rows|multiitem|textvariableeditor|multi line paste|readonly date|google signup|permission validation|draft email|pdf upload",
    re.I,
)
AI_RX = re.compile(r"\bai\b|assistant|composer|markdown|code-interpreter|record chips", re.I)
OPTIONAL_RX = re.compile(
    r"clear .*selection|disable going to settings|sync record selection|display found items|hide tabs for system objects|separate create draft cases",
    re.I,
)


def classify_batch(row: dict[str, str]) -> str:
    override = BATCH_OVERRIDES.get(row["short_sha"])
    if override:
        return override

    subject = row["subject"]
    top_dirs = row["top_dirs"]

    if SECURITY_RX.search(subject) or row["final_tag"] in {"security", "platform_fix"}:
        if SECURITY_DEP_RX.search(subject):
            return "1B"
        return "1A"

    if WORKFLOW_RX.search(subject) or row["final_tag"] == "workflow_fix":
        return "4A"

    if BACKEND_DATA_RX.search(subject):
        return "2B"

    if "packages/twenty-server" in top_dirs and "packages/twenty-front" not in top_dirs:
        return "2A"

    if OPTIONAL_RX.search(subject) or row["final_tag"] in {"ux_fix", "default_core_yes"}:
        return "5A"

    if AI_RX.search(subject):
        return "3C"

    if FRONT_FORM_RX.search(subject):
        return "3B"

    if FRONT_NAV_RX.search(subject):
        return "3A"

    return "3D"


def risk_level(row: dict[str, str], batch_code: str) -> str:
    files_changed = int(row["files_changed"])
    top_dirs = row["top_dirs"].count(",") + 1 if row["top_dirs"] else 0
    if batch_code.startswith("1"):
        return "high"
    if batch_code == "4A":
        return "medium"
    if files_changed >= 20 or top_dirs >= 3:
        return "high"
    if files_changed >= 6 or batch_code in {"2A", "2B", "5A"}:
        return "medium"
    return "low"


def build_plan_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    yes_rows = [row for row in rows if row["final_recommendation"] == "yes"]
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in yes_rows:
        batch_code = classify_batch(row)
        row["plan_batch"] = batch_code
        row["plan_wave"] = str(BATCH_INFO[batch_code]["wave"])
        row["plan_wave_name"] = BATCH_INFO[batch_code]["wave_name"]
        row["plan_batch_name"] = BATCH_INFO[batch_code]["batch_name"]
        row["plan_goal"] = BATCH_INFO[batch_code]["goal"]
        row["plan_validation"] = BATCH_INFO[batch_code]["validation"]
        row["plan_risk"] = risk_level(row, batch_code)
        grouped[batch_code].append(row)

    for batch_code, batch_rows in grouped.items():
        batch_rows.sort(key=lambda row: int(row["index"]))
        chunk_size = BATCH_INFO[batch_code]["chunk_size"]
        for offset, row in enumerate(batch_rows):
            chunk_num = offset // chunk_size + 1
            row["plan_chunk"] = f"{batch_code}-{chunk_num:02d}"
            override_chunk = CHUNK_OVERRIDES.get(row["short_sha"])
            if override_chunk:
                row["plan_chunk"] = override_chunk

    plan_rows: list[dict[str, str]] = []
    order = 1
    for batch_code in BATCH_ORDER:
        batch_rows = sorted(grouped.get(batch_code, []), key=lambda row: int(row["index"]))
        for row in batch_rows:
            plan_rows.append(
                {
                    "plan_order": str(order),
                    "wave": row["plan_wave"],
                    "wave_name": row["plan_wave_name"],
                    "batch": row["plan_batch"],
                    "batch_name": row["plan_batch_name"],
                    "chunk": row["plan_chunk"],
                    "risk": row["plan_risk"],
                    "review_index": row["index"],
                    "date": row["date"],
                    "short_sha": row["short_sha"],
                    "sha": row["sha"],
                    "subject": row["subject"],
                    "rationale": row["final_rationale"],
                    "top_dirs": row["top_dirs"],
                    "files_changed": row["files_changed"],
                    "goal": row["plan_goal"],
                    "validation": row["plan_validation"],
                    "commit_url": row["commit_url"],
                }
            )
            order += 1

    return plan_rows


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def build_chunk_rows(plan_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    chunks: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in plan_rows:
        chunks[row["chunk"]].append(row)

    chunk_rows = []
    for chunk in sorted(chunks.keys(), key=lambda value: (BATCH_ORDER.index(value.split("-")[0]), value)):
        chunk_items = chunks[chunk]
        batch_code = chunk.split("-")[0]
        info = BATCH_INFO[batch_code]
        chunk_rows.append(
            {
                "chunk": chunk,
                "wave": str(info["wave"]),
                "wave_name": info["wave_name"],
                "batch": batch_code,
                "batch_name": info["batch_name"],
                "commit_count": str(len(chunk_items)),
                "risk_mix": ", ".join(
                    f"{risk}:{count}" for risk, count in Counter(item["risk"] for item in chunk_items).most_common()
                ),
                "first_plan_order": chunk_items[0]["plan_order"],
                "last_plan_order": chunk_items[-1]["plan_order"],
                "validation": info["validation"],
                "cherry_pick_command": "git cherry-pick " + " ".join(item["sha"] for item in chunk_items),
            }
        )
    return chunk_rows


def write_xlsx(
    plan_rows: list[dict[str, str]],
    chunk_rows: list[dict[str, str]],
    review_rows: list[dict[str, str]],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    def fill_sheet(ws, rows):
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row[h] for h in headers])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.iter_cols():
            max_len = max(len(str(cell.value or "")) for cell in col[: min(ws.max_row, 80)])
            width = min(max(max_len + 2, 10), 70)
            ws.column_dimensions[col[0].column_letter].width = width
        for row in ws.iter_rows():
            for cell in row:
                if cell.column_letter in {"F", "L", "M", "P", "Q"}:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        return headers

    ws_plan = wb.active
    ws_plan.title = "Plan"
    fill_sheet(ws_plan, plan_rows)

    ws_chunks = wb.create_sheet("Chunk Summary")
    fill_sheet(ws_chunks, chunk_rows)

    ws_review = wb.create_sheet("Reviewed Yes")
    reviewed_yes = [row for row in review_rows if row["final_recommendation"] == "yes"]
    fill_sheet(ws_review, reviewed_yes)

    ws_watch = wb.create_sheet("Migration Watchlist")
    fill_sheet(ws_watch, WATCHLIST)

    ws_summary = wb.create_sheet("Summary")
    summary_rows = [
        {"metric": "Plan rows", "value": len(plan_rows)},
        {"metric": "Wave 1", "value": sum(1 for row in plan_rows if row["wave"] == "1")},
        {"metric": "Wave 2", "value": sum(1 for row in plan_rows if row["wave"] == "2")},
        {"metric": "Wave 3", "value": sum(1 for row in plan_rows if row["wave"] == "3")},
        {"metric": "Wave 4", "value": sum(1 for row in plan_rows if row["wave"] == "4")},
        {"metric": "Wave 5", "value": sum(1 for row in plan_rows if row["wave"] == "5")},
        {"metric": "Chunks", "value": len(chunk_rows)},
        {"metric": "Migration watchlist", "value": len(WATCHLIST)},
    ]
    fill_sheet(ws_summary, summary_rows)

    wb.save(PLAN_XLSX)


def main() -> None:
    review_rows = list(csv.DictReader(INPUT_CSV.open()))
    plan_rows = build_plan_rows(review_rows)
    chunk_rows = build_chunk_rows(plan_rows)
    write_csv(PLAN_CSV, plan_rows)
    write_xlsx(plan_rows, chunk_rows, review_rows)
    print(PLAN_CSV)
    print(f"wrote {len(plan_rows)} planned rows")
    print(PLAN_XLSX)


if __name__ == "__main__":
    main()
